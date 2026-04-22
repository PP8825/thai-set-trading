#!/usr/bin/env python3
"""
Thai SET — End-of-Day Portfolio Report
─────────────────────────────────────────────────────────────────
Runs at 4:35 PM Bangkok time after market close.
1. Loads live portfolio state from set_portfolio.json
2. Fetches current prices for all holdings
3. Generates a dated Excel report  →  Portfolio_Reports/Portfolio_YYYYMMDD.xlsx
4. Sends a detailed LINE message with the full portfolio table

Excel sheets
  Daily Summary   – portfolio value, cash, P&L today vs total
  Holdings        – every position: shares, avg cost, price, value, P&L
  Today's Trades  – BUY / SELL executed this session
  Trade History   – complete trade log since inception
"""

import sys, os, json, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── Auto-install ──────────────────────────────────────────────────────────────
def ensure_packages():
    import importlib, subprocess
    for pkg in ["yfinance", "pandas", "requests", "openpyxl"]:
        try:
            importlib.import_module(pkg)
        except ImportError:
            print(f"Installing {pkg}...")
            for args in [
                [sys.executable, "-m", "pip", "install", pkg, "-q"],
                [sys.executable, "-m", "pip", "install", pkg, "--user", "-q"],
                [sys.executable, "-m", "pip", "install", pkg,
                 "--break-system-packages", "-q"],
            ]:
                try:
                    subprocess.check_call(
                        args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    break
                except subprocess.CalledProcessError:
                    continue

ensure_packages()

import requests
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH    = os.path.join(SCRIPT_DIR, "set_config.json")
PORTFOLIO_PATH = os.path.join(SCRIPT_DIR, "set_portfolio.json")
REPORTS_DIR    = os.path.join(SCRIPT_DIR, "Portfolio_Reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

# ─── Config ────────────────────────────────────────────────────────────────────
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    cfg = json.load(f)

LINE_TOKEN   = os.environ.get("LINE_TOKEN", cfg.get("line_channel_access_token", ""))
LINE_USER_ID = os.environ.get("LINE_USER_ID", cfg.get("line_user_id", ""))

# ─── Style helpers (industry-standard colour coding) ──────────────────────────
# Blue  = hardcoded inputs  |  Black = formulas  |  Green = cross-sheet links
NAVY  = "1F3864"; MID  = "2E75B6"; LIGHT = "BDD7EE"
GRN   = "00B050"; RED  = "C00000"; YLW   = "FFC000"
LGRN  = "E2EFDA"; LRED = "FCE4D6"; GRY   = "D9D9D9"
WHITE = "FFFFFF"

def _font(bold=False, sz=10, color="000000"):
    return Font(name="Arial", bold=bold, size=sz, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _al(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _border():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)

def _apply(cell, bold=False, sz=10, color="000000", bg=None,
           h="center", wrap=False, fmt=None):
    cell.font      = _font(bold=bold, sz=sz, color=color)
    if bg:
        cell.fill  = _fill(bg)
    cell.alignment = _al(h=h, wrap=wrap)
    if fmt:
        cell.number_format = fmt
    cell.border    = _border()
    return cell

def _hdr(ws, row, col, val, bg=NAVY, fg=WHITE, sz=10, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=val)
    _apply(c, bold=bold, sz=sz, color=fg, bg=bg, h=h)
    return c

def _data(ws, row, col, val, fmt=None, color="000000", bg=None, bold=False, h="right"):
    c = ws.cell(row=row, column=col, value=val)
    _apply(c, bold=bold, sz=10, color=color, bg=bg, h=h, fmt=fmt)
    return c

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─── Price fetch ──────────────────────────────────────────────────────────────
def fetch_price(ticker):
    try:
        df = yf.Ticker(ticker).history(period="5d", auto_adjust=True)
        if not df.empty and "Close" in df.columns:
            return float(df["Close"].dropna().iloc[-1])
    except Exception:
        pass
    return None

def fetch_all_prices(tickers: list[str], workers: int = 8) -> dict:
    prices = {}
    with ThreadPoolExecutor(max_workers=workers) as ex:
        fut = {ex.submit(fetch_price, t): t for t in tickers}
        for f in as_completed(fut):
            t = fut[f]
            px = f.result()
            if px:
                prices[t] = px
    return prices

# ─── Portfolio helpers ────────────────────────────────────────────────────────
def portfolio_value(port: dict, prices: dict) -> float:
    val = port["cash"]
    for ticker, h in port["holdings"].items():
        val += h["shares"] * prices.get(ticker, h["avg_cost"])
    return val

def today_trades(port: dict) -> list:
    today_str = datetime.date.today().isoformat()
    return [t for t in port.get("trades", []) if str(t.get("date", "")) == today_str]

# ─── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(port: dict, prices: dict, report_date: datetime.date) -> str:
    from openpyxl import load_workbook

    MASTER_PATH = os.path.join(REPORTS_DIR, "Portfolio_Master.xlsx")

    # Load existing master file or create new one
    if os.path.exists(MASTER_PATH):
        wb = load_workbook(MASTER_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    total_val  = portfolio_value(port, prices)
    start_cap  = port["capital"]
    pnl_total  = total_val - start_cap
    pnl_pct    = pnl_total / start_cap * 100
    peak       = port.get("peak_value", start_cap)
    drawdown   = (total_val - peak) / peak * 100 if peak > 0 else 0
    day_no     = port.get("day_count", 1)
    start_date = port.get("start_date", report_date.isoformat())

    # ── traded today ──────────────────────────────────────────────────────────
    traded_today = today_trades(port)
    buy_today    = [t for t in traded_today if t["action"] == "BUY"]
    sell_today   = [t for t in traded_today if t["action"] == "SELL"]
    pnl_today    = sum(t.get("pnl", 0) for t in sell_today)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: DAILY PERFORMANCE (cumulative — one row per day)
    # ═══════════════════════════════════════════════════════════════════════
    if "Daily Performance" in wb.sheetnames:
        ws1 = wb["Daily Performance"]
    else:
        ws1 = wb.create_sheet("Daily Performance")
    ws1.sheet_view.showGridLines = False
    _set_col_widths(ws1, [12, 6, 16, 14, 16, 14, 12, 10, 10, 10, 8, 8])

    # Header row (only if sheet is new)
    if ws1.max_row <= 1 and ws1["A1"].value is None:
        hdrs = ["Date", "Day", "Portfolio Value (฿)", "Cash (฿)",
                "Invested (฿)", "Total P&L (฿)", "Total Return %",
                "Realised P&L Today (฿)", "Drawdown %",
                "Positions", "Buys", "Sells"]
        for c, h in enumerate(hdrs, 1):
            _hdr(ws1, 1, c, h, bg=NAVY, sz=10)
        ws1.row_dimensions[1].height = 22

    # Check if today already has a row — update it if so
    today_str = report_date.isoformat()
    existing_row = None
    for row in ws1.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value == today_str:
            existing_row = row[0].row
            break

    new_row_data = [
        today_str,
        day_no,
        total_val,
        port["cash"],
        total_val - port["cash"],
        pnl_total,
        pnl_pct / 100,
        pnl_today,
        drawdown / 100,
        len(port["holdings"]),
        len(buy_today),
        len(sell_today),
    ]
    fmts = [
        "YYYY-MM-DD", "#,##0", "฿#,##0", "฿#,##0", "฿#,##0",
        "฿#,##0;(฿#,##0);-", "0.00%", "฿#,##0;(฿#,##0);-",
        "0.00%", "#,##0", "#,##0", "#,##0"
    ]
    colors = [
        "000000", "595959",
        GRN if total_val > start_cap else RED,
        MID, "595959",
        GRN if pnl_total >= 0 else RED,
        GRN if pnl_pct >= 0 else RED,
        GRN if pnl_today >= 0 else RED,
        RED if drawdown < -3 else "595959",
        MID, GRN, RED
    ]

    target_row = existing_row if existing_row else ws1.max_row + 1
    bg = LGRN if pnl_total >= 0 else LRED
    for c, (val, fmt, col) in enumerate(zip(new_row_data, fmts, colors), 1):
        cell = ws1.cell(target_row, c, val)
        _apply(cell, color=col, bg=bg if c > 1 else WHITE, fmt=fmt,
               bold=(c in [3, 6, 7]), h="center" if c != 1 else "left")
    ws1.row_dimensions[target_row].height = 20

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: HOLDINGS (replaced each day with current state)
    # ═══════════════════════════════════════════════════════════════════════
    if "Holdings" in wb.sheetnames:
        del wb["Holdings"]
    ws2 = wb.create_sheet("Holdings")
    ws2.sheet_view.showGridLines = False
    _set_col_widths(ws2, [12, 14, 8, 13, 13, 14, 18, 16, 14])

    ws2.merge_cells("A1:I1")
    t2 = ws2["A1"]
    t2.value     = f"Current Holdings  —  {report_date.strftime('%d %b %Y')}"
    t2.font      = Font(name="Arial", bold=True, size=14, color="C9A84C")
    t2.fill      = _fill(NAVY)
    t2.alignment = _al()
    ws2.row_dimensions[1].height = 34

    h2_labels = ["Ticker", "Stock Name", "Shares", "Avg Cost (฿)",
                 "Current Price (฿)", "Market Value (฿)",
                 "Unrealised P&L (฿)", "P&L %", "Entry Date"]
    for c, h in enumerate(h2_labels, 1):
        _hdr(ws2, 2, c, h, bg=MID, sz=9)
    ws2.row_dimensions[2].height = 22

    sorted_holdings = sorted(
        port["holdings"].items(),
        key=lambda x: -(x[1]["shares"] * prices.get(x[0], x[1]["avg_cost"]))
    )

    total_mkt_val  = 0
    total_unreal   = 0
    for i, (ticker, h) in enumerate(sorted_holdings):
        r      = i + 3
        px     = prices.get(ticker, h["avg_cost"])
        mkt    = h["shares"] * px
        cost   = h["shares"] * h["avg_cost"]
        unreal = mkt - cost
        pnl_p  = unreal / cost if cost > 0 else 0
        total_mkt_val  += mkt
        total_unreal   += unreal

        bg_row  = LGRN if pnl_p > 0 else (LRED if pnl_p < -0.02 else WHITE)
        pnl_col = GRN  if unreal >= 0 else RED

        _data(ws2, r, 1, ticker,          h="left",   bg=bg_row, bold=True, color=NAVY)
        _data(ws2, r, 2, h["name"],       h="left",   bg=bg_row)
        _data(ws2, r, 3, h["shares"],     fmt="#,##0",bg=bg_row)
        # Avg cost = blue (hardcoded input from purchase price)
        _data(ws2, r, 4, h["avg_cost"],   fmt="฿#,##0.00;(฿#,##0.00);-",
              bg=bg_row, color="0000FF")
        # Current price = black (fetched/formula)
        _data(ws2, r, 5, px,              fmt="฿#,##0.00;(฿#,##0.00);-", bg=bg_row)
        # Market value = formula (black text)
        mv = ws2.cell(r, 6, f"=C{r}*E{r}")
        _apply(mv, fmt="฿#,##0;(฿#,##0);-", bg=bg_row)
        # Unrealised P&L = formula
        up = ws2.cell(r, 7, f"=F{r}-C{r}*D{r}")
        _apply(up, color=pnl_col, bold=True, fmt='฿#,##0;(฿#,##0);"-"', bg=bg_row)
        # P&L % = formula
        pp = ws2.cell(r, 8, f"=(E{r}-D{r})/D{r}")
        _apply(pp, color=pnl_col, fmt='0.00%;(0.00%);"-"', bg=bg_row)
        _data(ws2, r, 9, h.get("entry_date", ""),  h="center", bg=bg_row, color="595959")
        ws2.row_dimensions[r].height = 20

    # Totals
    tot_row = len(sorted_holdings) + 3
    ws2.row_dimensions[tot_row].height = 22
    _data(ws2, tot_row, 2, "TOTAL", bold=True, h="right")
    mv_tot = ws2.cell(tot_row, 6, f"=SUM(F3:F{tot_row-1})")
    _apply(mv_tot, bold=True, fmt="฿#,##0;(฿#,##0);-")
    up_tot = ws2.cell(tot_row, 7, f"=SUM(G3:G{tot_row-1})")
    up_col = GRN if total_unreal >= 0 else RED
    _apply(up_tot, bold=True, color=up_col, fmt='฿#,##0;(฿#,##0);"-"')
    pp_tot = ws2.cell(tot_row, 8, f"=G{tot_row}/SUMPRODUCT(C3:C{tot_row-1},D3:D{tot_row-1})")
    _apply(pp_tot, bold=True, color=up_col, fmt='0.00%;(0.00%);"-"')

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3: TODAY'S TRADES (replaced each day)
    # ═══════════════════════════════════════════════════════════════════════
    if "Today's Trades" in wb.sheetnames:
        del wb["Today's Trades"]
    ws3 = wb.create_sheet("Today's Trades")
    ws3.sheet_view.showGridLines = False
    _set_col_widths(ws3, [13, 14, 8, 8, 13, 8, 14, 13, 18])

    ws3.merge_cells("A1:I1")
    t3 = ws3["A1"]
    t3.value     = f"Today's Trades  —  {report_date.strftime('%d %b %Y')}"
    t3.font      = Font(name="Arial", bold=True, size=14, color="C9A84C")
    t3.fill      = _fill(NAVY)
    t3.alignment = _al()
    ws3.row_dimensions[1].height = 34

    if traded_today:
        h3 = ["Date","Stock","Shares","Action","Price (฿)",
              "Lots","Value (฿)","Realised P&L (฿)","Reason / Signal"]
        for c, h in enumerate(h3, 1):
            _hdr(ws3, 2, c, h, bg=MID, sz=9)
        ws3.row_dimensions[2].height = 22

        for i, t in enumerate(traded_today):
            r      = i + 3
            action = t.get("action", "")
            is_buy = action == "BUY"
            bg_row = LGRN if is_buy else (LRED if t.get("pnl", 0) < 0 else "FFF9C4")
            pnl_c  = GRN  if t.get("pnl", 0) >= 0 else RED

            _data(ws3, r, 1, str(t.get("date", "")),    h="center", bg=bg_row)
            _data(ws3, r, 2, t.get("name", ""),          h="left",   bg=bg_row, bold=True)
            _data(ws3, r, 3, t.get("shares", 0),         fmt="#,##0",bg=bg_row)
            c4 = ws3.cell(r, 4, action)
            _apply(c4, bold=True, sz=10,
                   color=GRN if is_buy else RED,
                   bg=LGRN if is_buy else LRED)
            _data(ws3, r, 5, t.get("price", 0),
                  fmt="฿#,##0.00;(฿#,##0.00);-",        bg=bg_row, color="0000FF")
            lots = t.get("shares", 0) // 100
            _data(ws3, r, 6, lots,                        fmt="#,##0", bg=bg_row)
            val_c = ws3.cell(r, 7, f"=C{r}*E{r}")
            _apply(val_c, fmt="฿#,##0;(฿#,##0);-",       bg=bg_row)
            pnl_v = t.get("pnl", 0) if not is_buy else None
            _data(ws3, r, 8, pnl_v if pnl_v is not None else "—",
                  fmt='฿#,##0;(฿#,##0);"-"',             bg=bg_row,
                  color=pnl_c if pnl_v else "595959")
            _data(ws3, r, 9, t.get("reason", ""),         h="left",
                  color="595959",                          bg=bg_row)
            ws3.row_dimensions[r].height = 20

        # Total P&L row
        bot = len(traded_today) + 3
        ws3.row_dimensions[bot].height = 22
        _data(ws3, bot, 7, "TODAY'S P&L", bold=True, h="right")
        tp = ws3.cell(bot, 8, f"=SUM(H3:H{bot-1})")
        _apply(tp, bold=True, sz=11, color=GRN if pnl_today >= 0 else RED,
               fmt='฿#,##0;(฿#,##0);"-"')
    else:
        ws3.merge_cells("A3:I3")
        ws3["A3"] = "No trades were executed today."
        ws3["A3"].font      = Font(name="Arial", size=11, italic=True, color="595959")
        ws3["A3"].alignment = _al()
        ws3.row_dimensions[3].height = 30

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4: TRADE HISTORY (cumulative — all trades ever)
    # ═══════════════════════════════════════════════════════════════════════
    if "Trade History" in wb.sheetnames:
        del wb["Trade History"]
    ws4 = wb.create_sheet("Trade History")
    ws4.sheet_view.showGridLines = False
    _set_col_widths(ws4, [13, 12, 14, 8, 8, 13, 10, 14, 14, 18])

    ws4.merge_cells("A1:J1")
    t4 = ws4["A1"]
    t4.value     = f"Complete Trade History  —  since {start_date}"
    t4.font      = Font(name="Arial", bold=True, size=14, color="C9A84C")
    t4.fill      = _fill(NAVY)
    t4.alignment = _al()
    ws4.row_dimensions[1].height = 34

    all_trades = port.get("trades", [])
    if all_trades:
        h4 = ["Date","Ticker","Stock","Action","Shares","Price (฿)",
              "Lots","Value (฿)","Realised P&L (฿)","Reason / Signal"]
        for c, h in enumerate(h4, 1):
            _hdr(ws4, 2, c, h, bg=MID, sz=9)
        ws4.row_dimensions[2].height = 22

        for i, t in enumerate(sorted(all_trades, key=lambda x: str(x.get("date",""))), 3):
            action = t.get("action", "")
            is_buy = action == "BUY"
            bg_row = LIGHT if i % 2 == 1 else WHITE
            pnl_v  = t.get("pnl")

            _data(ws4, i, 1,  str(t.get("date", "")),   h="center",  bg=bg_row)
            _data(ws4, i, 2,  t.get("ticker", ""),       h="center",  bg=bg_row, bold=True)
            _data(ws4, i, 3,  t.get("name", ""),         h="left",    bg=bg_row)
            c4 = ws4.cell(i, 4, action)
            _apply(c4, bold=True, sz=9,
                   color=GRN if is_buy else RED,
                   bg=LGRN if is_buy else LRED)
            _data(ws4, i, 5,  t.get("shares", 0),        fmt="#,##0", bg=bg_row)
            _data(ws4, i, 6,  t.get("price",  0),
                  fmt="฿#,##0.00;(฿#,##0.00);-",         bg=bg_row, color="0000FF")
            _data(ws4, i, 7,  t.get("shares", 0) // 100, fmt="#,##0",bg=bg_row)
            val = ws4.cell(i, 8, f"=E{i}*F{i}")
            _apply(val, fmt="฿#,##0;(฿#,##0);-",          bg=bg_row)
            if pnl_v is not None and not is_buy:
                _data(ws4, i, 9, pnl_v, fmt='฿#,##0;(฿#,##0);"-"',
                      bg=bg_row, color=GRN if pnl_v >= 0 else RED)
            else:
                _data(ws4, i, 9, "—", bg=bg_row, color="595959")
            _data(ws4, i, 10, t.get("reason", ""),  h="left", color="595959", bg=bg_row)
            ws4.row_dimensions[i].height = 17

        # Summary totals
        last = len(all_trades) + 2
        tot  = last + 2
        ws4.row_dimensions[tot].height = 22
        for c, label, col_letter in [
            (7, "TOTAL TRADES", None),
            (8, None, "H"), (9, "TOTAL P&L", "I")
        ]:
            if label:
                ws4.cell(tot, c, label).font = _font(bold=True, sz=10)
            if col_letter:
                f = ws4.cell(tot, c, f"=SUM({col_letter}3:{col_letter}{last})")
                f.number_format = "฿#,##0;(฿#,##0);-"
                f.font          = _font(bold=True, sz=11, color=GRN)
                f.alignment     = _al()
                f.border        = _border()
    else:
        ws4.merge_cells("A3:J3")
        ws4["A3"] = "No trades recorded yet."
        ws4["A3"].font      = Font(name="Arial", size=11, italic=True, color="595959")
        ws4["A3"].alignment = _al()

    # ── Tab colours & order ───────────────────────────────────────────────────
    wb["Daily Performance"].sheet_properties.tabColor = NAVY
    wb["Holdings"].sheet_properties.tabColor          = GRN
    wb["Today's Trades"].sheet_properties.tabColor    = "ED7D31"
    wb["Trade History"].sheet_properties.tabColor     = MID

    # Always save as master file
    wb.save(MASTER_PATH)
    return MASTER_PATH

# ─── LINE message ─────────────────────────────────────────────────────────────
def build_eod_message(port: dict, prices: dict,
                      report_date: datetime.date, excel_path: str) -> str:
    total_val    = portfolio_value(port, prices)
    start_cap    = port["capital"]
    pnl_total    = total_val - start_cap
    pnl_pct      = pnl_total / start_cap * 100
    peak         = port.get("peak_value", start_cap)
    drawdown     = (total_val - peak) / peak * 100 if peak > 0 else 0
    traded_today = today_trades(port)
    pnl_today    = sum(t.get("pnl", 0) for t in traded_today if t["action"] == "SELL")

    p_sign = "+" if pnl_total >= 0 else ""
    t_sign = "+" if pnl_today >= 0 else ""
    p_icon = "📈" if pnl_total >= 0 else "📉"
    dd_str = f"  ⚠️ Drawdown {drawdown:.2f}%" if drawdown < -3 else f"  Drawdown {drawdown:.2f}%"

    lines = [
        f"🇹🇭 SET Portfolio — End of Day",
        f"📅 {report_date.strftime('%a %d %b %Y')} | Day {port.get('day_count',1)}",
        f"Since {port.get('start_date','—')}",
        "─" * 34,
        "",
        "💼 PORTFOLIO",
        f"  Value    : ฿{total_val:>11,.0f}",
        f"  Cash     : ฿{port['cash']:>11,.0f}",
        f"  Invested : ฿{(total_val-port['cash']):>11,.0f}",
        f"  Total P&L: {p_icon} {p_sign}฿{pnl_total:,.0f} ({p_sign}{pnl_pct:.2f}%)",
        f"{dd_str}",
        "",
    ]

    # Today's trades
    if traded_today:
        buy_t  = [t for t in traded_today if t["action"] == "BUY"]
        sell_t = [t for t in traded_today if t["action"] == "SELL"]
        lines.append(f"📋 TRADES TODAY  ({len(traded_today)} orders)")
        for t in buy_t:
            lines.append(f"  🟢 BUY   {t['name']:8s}  {t['shares']:,d}sh"
                         f" @ ฿{t['price']:,.2f} = ฿{t['value']:,.0f}")
        for t in sell_t:
            ps = "+" if t.get('pnl',0) >= 0 else ""
            lines.append(f"  🔴 SELL  {t['name']:8s}  {t['shares']:,d}sh"
                         f" @ ฿{t['price']:,.2f} = ฿{t['value']:,.0f}"
                         f"  ({ps}฿{t.get('pnl',0):,.0f})")
        lines.append(f"  Today's realised P&L: {t_sign}฿{pnl_today:,.0f}")
    else:
        lines.append("📋 TRADES TODAY  No transactions")
    lines.append("")

    # Holdings table
    if port["holdings"]:
        lines.append(f"📊 HOLDINGS  ({len(port['holdings'])} stocks)")
        sorted_h = sorted(
            port["holdings"].items(),
            key=lambda x: -(x[1]["shares"] * prices.get(x[0], x[1]["avg_cost"]))
        )
        for i, (ticker, h) in enumerate(sorted_h, 1):
            px     = prices.get(ticker, h["avg_cost"])
            mkt    = h["shares"] * px
            cost   = h["shares"] * h["avg_cost"]
            unreal = mkt - cost
            u_pct  = unreal / cost * 100 if cost > 0 else 0
            u_icon = "▲" if unreal >= 0 else "▼"
            u_sign = "+" if unreal >= 0 else ""
            lines.append(
                f"  {i:2d}. {h['name']:8s}  {h['shares']:5,d}sh"
                f"  ฿{h['avg_cost']:,.2f}→฿{px:,.2f}"
                f"  {u_icon}{u_sign}{u_pct:.1f}%  ({u_sign}฿{unreal:,.0f})"
            )
        total_unreal = sum(
            (h["shares"] * prices.get(t, h["avg_cost"])) - (h["shares"] * h["avg_cost"])
            for t, h in port["holdings"].items()
        )
        u_sign = "+" if total_unreal >= 0 else ""
        lines.append(f"  ─────────────────────────────────")
        lines.append(f"  Total unrealised: {u_sign}฿{total_unreal:,.0f}")
    else:
        lines.append("📊 HOLDINGS  No open positions")

    lines += [
        "",
        f"📁 Master file updated: Portfolio_Master.xlsx",
        f"   (All daily history in one file)",
        "",
        "─" * 34,
        "⚠️ Educational only. Not financial advice.",
    ]
    return "\n".join(lines)

# ─── LINE API ─────────────────────────────────────────────────────────────────
def send_line(message: str) -> tuple:
    try:
        resp = requests.post(
            "https://api.line.me/v2/bot/message/push",
            headers={"Authorization": f"Bearer {LINE_TOKEN}",
                     "Content-Type": "application/json"},
            data=json.dumps({"to": LINE_USER_ID,
                             "messages": [{"type": "text", "text": message}]},
                            ensure_ascii=False).encode("utf-8"),
            timeout=15,
        )
        return resp.status_code == 200, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return False, str(e)

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    today = datetime.date.today()

    print("=" * 56)
    print("Thai SET — End-of-Day Portfolio Report")
    print(f"Run: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 56)

    # Load portfolio
    if not os.path.exists(PORTFOLIO_PATH):
        print("⚠️  No portfolio file found. Run set_signal_alert.py first.")
        sys.exit(1)

    with open(PORTFOLIO_PATH, "r", encoding="utf-8") as f:
        port = json.load(f)

    if port["day_count"] == 0:
        print("⚠️  Portfolio not started yet. Run set_signal_alert.py first.")
        sys.exit(0)

    # Fetch current prices for holdings + any traded-today tickers
    all_tickers = list(port["holdings"].keys())
    today_t = today_trades(port)
    all_tickers += [t["ticker"] for t in today_t if t["ticker"] not in all_tickers]
    all_tickers = list(dict.fromkeys(all_tickers))  # deduplicate

    if all_tickers:
        print(f"\nFetching prices for {len(all_tickers)} tickers...")
        prices = fetch_all_prices(all_tickers)
        print(f"  Got prices for {len(prices)}/{len(all_tickers)}")
    else:
        prices = {}

    # Fill in avg_cost for any tickers without live price
    for ticker, h in port["holdings"].items():
        if ticker not in prices:
            prices[ticker] = h["avg_cost"]
            print(f"  ⚠️  Using avg cost for {ticker} (no live price)")

    # Increment day count
    port["day_count"] = port.get("day_count", 0) + 1
    print(f"  Day count updated to: {port['day_count']}")

    # Update peak value
    total_val_now = port["cash"] + sum(
        h["shares"] * prices.get(t, h["avg_cost"])
        for t, h in port["holdings"].items()
    )
    if total_val_now > port.get("peak_value", port["capital"]):
        port["peak_value"] = total_val_now

    # Save updated portfolio
    with open(PORTFOLIO_PATH, "w", encoding="utf-8") as f:
        json.dump(port, f, indent=2, ensure_ascii=False)

    # Build Excel
    print("\nGenerating Excel report...")
    excel_path = build_excel(port, prices, today)
    print(f"  Saved: {excel_path}")

    # Build and send LINE message
    message = build_eod_message(port, prices, today, excel_path)
    print("\n── LINE Message Preview ─────────────────────────")
    print(message)

    print("\n── Sending to LINE... ───────────────────────────")
    ok, resp = send_line(message)
    if ok:
        print("✅ Sent successfully!")
    else:
        print(f"❌ Failed: {resp}")
        sys.exit(1)

    print(f"\n📁 Excel saved to: {excel_path}")
    print("Done.")

if __name__ == "__main__":
    main()
