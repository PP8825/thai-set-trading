#!/usr/bin/env python3
"""
Thai SET Backtest & Forward Test  |  RSI / SMA50 / MACD Strategy
═══════════════════════════════════════════════════════════════════
Backtest : 1 year (Apr 2025 → Apr 2026)   Signal lag : 1 day (no look-ahead)
Forward  : Current signals → recommended positions for next 1 month
Output   : SET_Backtest_Report.xlsx  in the same folder as this script

Strategy
────────
• Long-only, equal-weight daily rebalance
• BUY  when score ≥ +1  (≥2 of 3 indicators bullish)
• FLAT when score ≤  0  (no position)
• RSI(14) < 30 → +1 | RSI > 70 → -1 | else 0
• Close > SMA50 → +1 | else -1
• MACD line > Signal line → +1 | else -1
• Transaction cost: 0.25% per side  |  Starting capital: ฿1,000,000
"""

import sys, os, json, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── Auto-install ─────────────────────────────────────────────────────────────
def ensure_pkgs():
    import importlib, subprocess
    for pkg in ["yfinance","pandas","numpy","openpyxl","requests"]:
        try: importlib.import_module(pkg)
        except ImportError:
            print(f"Installing {pkg}...")
            for args in [
                [sys.executable,"-m","pip","install",pkg,"-q"],
                [sys.executable,"-m","pip","install",pkg,"--user","-q"],
            ]:
                try: subprocess.check_call(args,stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL); break
                except: continue
ensure_pkgs()

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule

# ─── Config ───────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH  = os.path.join(SCRIPT_DIR, "set_config.json")
OUTPUT_PATH  = os.path.join(SCRIPT_DIR, "SET_Backtest_Report.xlsx")

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    cfg = json.load(f)

INSTRUMENTS  = [(i["name"], i["ticker"]) for i in cfg.get("instruments", [])]
BENCHMARK    = ("SET Index", "^SET.BK")

CAPITAL      = 1_000_000   # ฿ — used for historical backtest (fair comparison)
FWD_CAPITAL  = 300_000     # ฿ — live forward portfolio capital
TX_COST      = 0.0025      # 0.25% per side
RISK_FREE    = 0.015       # 1.5% p.a. (Thai deposit rate)
RSI_N        = cfg.get("rsi_period", 14)
RSI_OB       = cfg.get("rsi_overbought", 70)
RSI_OS       = cfg.get("rsi_oversold", 30)
SMA_N        = cfg.get("sma_period", 50)
BT_DAYS      = 365
WARMUP_DAYS  = 120         # extra history for indicator warmup
FWRD_DAYS    = 22          # ~1 trading month
FWD_MAX_POS  = 10          # max simultaneous holdings in forward portfolio
FWD_LOT      = 100         # Thai standard lot size
STOP_LOSS    = 0.08        # 8% stop-loss on forward positions

TODAY        = datetime.date.today()
BT_END       = pd.Timestamp(TODAY)
BT_START     = BT_END - pd.DateOffset(days=BT_DAYS)
FETCH_PERIOD = f"{BT_DAYS + WARMUP_DAYS}d"

# ─── Style helpers ────────────────────────────────────────────────────────────
def F(bold=False, sz=10, color="000000", name="Arial"):
    return Font(name=name, bold=bold, size=sz, color=color)
def Fill(c): return PatternFill("solid", fgColor=c)
def Al(h="center", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def Brd():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)
def apply(cell, bold=False, sz=10, color="000000", bg=None, h="center", wrap=False, fmt=None):
    cell.font = F(bold=bold, sz=sz, color=color)
    if bg: cell.fill = Fill(bg)
    cell.alignment = Al(h=h, wrap=wrap)
    if fmt: cell.number_format = fmt
    cell.border = Brd()

# ─── Indicators (vectorized, no look-ahead) ───────────────────────────────────
def rsi(s, n=RSI_N):
    d=s.diff(); ag=d.clip(lower=0).ewm(alpha=1/n,min_periods=n,adjust=False).mean()
    al=(-d).clip(lower=0).ewm(alpha=1/n,min_periods=n,adjust=False).mean()
    return 100-100/(1+ag/al.replace(0,1e-10))

def add_signals(close: pd.Series, open_: pd.Series) -> pd.DataFrame:
    df = pd.DataFrame({"Close": close, "Open": open_})
    df["RSI"]     = rsi(df["Close"])
    df["SMA50"]   = df["Close"].rolling(SMA_N).mean()
    ml            = df["Close"].ewm(span=12,adjust=False).mean() - df["Close"].ewm(span=26,adjust=False).mean()
    sl            = ml.ewm(span=9,adjust=False).mean()
    df["MACD"]    = ml; df["MACDSig"] = sl

    # Shift indicators by 1 day → signal generated at yesterday's close
    r = df["RSI"].shift(1)
    df["RSI_sig"]  = r.apply(lambda x: 1 if pd.notna(x) and x<RSI_OS else (-1 if pd.notna(x) and x>RSI_OB else 0))
    cp, sp         = df["Close"].shift(1), df["SMA50"].shift(1)
    df["MA_sig"]   = np.where(cp.notna()&sp.notna(), np.where(cp>sp,1,-1), 0)
    mp, sp2        = df["MACD"].shift(1), df["MACDSig"].shift(1)
    df["MACD_sig"] = np.where(mp.notna()&sp2.notna(), np.where(mp>sp2,1,-1), 0)
    df["Score"]    = df["RSI_sig"]+df["MA_sig"]+df["MACD_sig"]
    df["Long"]     = (df["Score"]>=1).astype(int)
    return df

# ─── Data fetch ───────────────────────────────────────────────────────────────
def fetch(name, ticker):
    try:
        df = yf.download(ticker, period=FETCH_PERIOD, auto_adjust=True, progress=False, timeout=30)
        if isinstance(df.columns, pd.MultiIndex): df.columns = df.columns.get_level_values(0)
        if df is None or df.empty or len(df)<60: return name, ticker, None
        df.index = pd.to_datetime(df.index).tz_localize(None)
        return name, ticker, df
    except Exception as e:
        return name, ticker, None

# ─── Portfolio backtest ───────────────────────────────────────────────────────
def backtest(stock_dict):
    """
    stock_dict: {ticker: (name, signals_df)}
    Returns (equity_df, trade_list, per_stock_list)
    """
    all_dates = sorted(set.union(*[set(df.index) for _,df in stock_dict.values()]))
    capital   = CAPITAL
    positions = {}   # {ticker: shares}
    equity    = []
    trades    = []

    for date in all_dates:
        if date < BT_START: continue

        # Target: stocks with Long=1 signal today
        target = {t: n for t,(n,df) in stock_dict.items()
                  if date in df.index and df.loc[date,"Long"]==1}

        # Value positions at today's open
        pv = capital
        for t, sh in positions.items():
            _, df = stock_dict[t]
            if date in df.index and pd.notna(df.loc[date,"Open"]):
                pv += sh * df.loc[date,"Open"]

        # Sell exits
        for t in [t for t in positions if t not in target]:
            _, df = stock_dict[t]
            if date in df.index and pd.notna(df.loc[date,"Open"]) and positions[t]>0:
                px = df.loc[date,"Open"]; sh = positions.pop(t)
                net = sh*px*(1-TX_COST); capital += net
                trades.append({"Date":date.date(),"Ticker":t,"Name":stock_dict[t][0],
                               "Action":"SELL","Price":round(px,2),"Shares":sh,
                               "Value":round(sh*px,0),"TxCost":round(sh*px*TX_COST,0)})

        # Buy entries
        n_new = len([t for t in target if t not in positions])
        if n_new>0 and capital>0:
            alloc = pv / max(len(target),1)
            for t,n in target.items():
                if t not in positions:
                    _, df = stock_dict[t]
                    if date in df.index and pd.notna(df.loc[date,"Open"]):
                        px = df.loc[date,"Open"]
                        sh = int(alloc/px/100)*100
                        if sh>0 and capital >= sh*px*(1+TX_COST):
                            capital -= sh*px*(1+TX_COST)
                            positions[t] = sh
                            trades.append({"Date":date.date(),"Ticker":t,"Name":n,
                                           "Action":"BUY","Price":round(px,2),"Shares":sh,
                                           "Value":round(sh*px,0),"TxCost":round(sh*px*TX_COST,0)})

        # Mark at close
        eod = capital
        for t,sh in positions.items():
            _, df = stock_dict[t]
            if date in df.index and pd.notna(df.loc[date,"Close"]): eod += sh*df.loc[date,"Close"]
        equity.append({"Date":date, "Portfolio":round(eod,0)})

    eq_df = pd.DataFrame(equity).set_index("Date")
    # Per-stock summary
    per_stock = []
    for t,(n,df) in stock_dict.items():
        df_bt = df[df.index>=BT_START]
        if len(df_bt)<5: continue
        st_trades = [x for x in trades if x["Ticker"]==t]
        buys  = [x for x in st_trades if x["Action"]=="BUY"]
        sells = [x for x in st_trades if x["Action"]=="SELL"]
        wins  = sum(1 for b,s in zip(buys,sells) if s["Price"]>b["Price"])
        bh    = (float(df_bt["Close"].iloc[-1])/float(df_bt["Close"].iloc[0])-1)*100
        last  = df.iloc[-1]
        sc    = int(last["Score"]) if pd.notna(last["Score"]) else 0
        ov    = ("STRONG BUY" if sc>=2 else "BUY" if sc==1 else
                 "HOLD" if sc==0 else "SELL" if sc>=-1 else "STRONG SELL")
        per_stock.append({
            "Ticker":t,"Name":n,
            "Buy-Hold (%)":round(bh,1),
            "# Trades":len(buys),
            "Wins":wins,"Losses":max(0,len(sells)-wins),
            "Win Rate (%)":round(wins/len(sells)*100,1) if sells else 0,
            "Current Signal":ov,"Score":sc,
            "Price (฿)":round(float(df["Close"].iloc[-1]),2),
            "RSI":round(float(df["RSI"].iloc[-1]),1) if pd.notna(df["RSI"].iloc[-1]) else None,
            "SMA50 (฿)":round(float(df["SMA50"].iloc[-1]),2) if pd.notna(df["SMA50"].iloc[-1]) else None,
        })
    return eq_df, pd.DataFrame(trades), pd.DataFrame(per_stock)

# ─── Performance metrics ──────────────────────────────────────────────────────
def perf(s: pd.Series, label="Strategy"):
    if len(s)<2: return {}
    ret  = s.pct_change().dropna()
    tot  = (s.iloc[-1]/s.iloc[0]-1)*100
    n    = len(ret)
    ann  = ((1+tot/100)**(252/n)-1)*100 if n>0 else 0
    vol  = float(ret.std()*252**0.5*100)
    sh   = (ann/100-RISK_FREE)/(vol/100) if vol>0 else 0
    dd   = (s-s.cummax())/s.cummax()*100
    return {"Label":label,"Total Return (%)":round(tot,2),"Annualised (%)":round(ann,2),
            "Volatility (%)":round(vol,2),"Sharpe Ratio":round(sh,2),
            "Max Drawdown (%)":round(float(dd.min()),2),
            "Final Value (฿)":int(s.iloc[-1]),"P&L (฿)":int(s.iloc[-1]-s.iloc[0])}

# ─── Forward test ─────────────────────────────────────────────────────────────
def forward_test(stock_dict):
    """
    Returns (signal_df, portfolio_df, tracker_df)
    signal_df    — all stocks with current signals
    portfolio_df — initial ฿300k portfolio: top-10 BUY allocations
    tracker_df   — 22-day walk-forward simulation using last 22 days of data
    """
    rows = []
    for t,(n,df) in stock_dict.items():
        if len(df)<5: continue
        last   = df.iloc[-1]
        sc     = int(last["Score"]) if pd.notna(last["Score"]) else 0
        ov     = ("STRONG BUY" if sc>=2 else "BUY" if sc==1 else
                  "HOLD" if sc==0 else "SELL" if sc>=-1 else "STRONG SELL")
        price  = float(last["Close"])
        sma    = float(last["SMA50"]) if pd.notna(last["SMA50"]) else None
        rsi_v  = float(last["RSI"])   if pd.notna(last["RSI"])   else None
        macd_v = float(last["MACD"])  if pd.notna(last["MACD"])  else None
        target = round(price*1.03,2) if sc>=1 else (round(price*0.97,2) if sc<=-1 else None)
        stop   = round(price*(1-STOP_LOSS),2) if sc>=1 else None
        action = "HOLD" if sc==0 else ("BUY" if sc>=1 else "AVOID/EXIT")
        rows.append({
            "Ticker":t,"Name":n,"Action":action,"Signal":ov,"Score":sc,
            "Current Price (฿)":round(price,2),
            "SMA50 (฿)":round(sma,2) if sma else None,
            "RSI":round(rsi_v,1) if rsi_v else None,
            "MACD":round(macd_v,4) if macd_v else None,
            "1-Month Target (฿)":target,
            "Stop-Loss (฿)":stop,
            "Upside (%)":round((target/price-1)*100,1) if target and sc>=1 else None,
        })
    signal_df = pd.DataFrame(rows).sort_values("Score",ascending=False).reset_index(drop=True)

    # ── Initial ฿300k portfolio allocation ─────────────────────────────────────
    cash_floor = FWD_CAPITAL * 0.05
    buy_rows   = signal_df[signal_df["Action"]=="BUY"].head(FWD_MAX_POS).copy()
    n_pos      = len(buy_rows)
    alloc_each = (FWD_CAPITAL - cash_floor) / n_pos if n_pos else 0
    port_rows  = []
    total_invested = 0
    for _, row in buy_rows.iterrows():
        px     = row["Current Price (฿)"]
        if not px or px <= 0: continue
        shares = int(alloc_each / px / FWD_LOT) * FWD_LOT
        if shares <= 0: shares = FWD_LOT
        cost   = shares * px * (1 + TX_COST)
        total_invested += cost
        port_rows.append({
            "Ticker":        row["Ticker"],
            "Name":          row["Name"],
            "Signal":        row["Signal"],
            "Score":         row["Score"],
            "Entry Price(฿)":round(px,2),
            "Shares":        shares,
            "Invested (฿)":  round(cost,0),
            "Weight (%)":    round(cost/FWD_CAPITAL*100,1),
            "1M Target (฿)": row["1-Month Target (฿)"],
            "Stop-Loss (฿)": row["Stop-Loss (฿)"],
            "Target P&L (฿)":round((row["1-Month Target (฿)"]-px)*shares,0) if row["1-Month Target (฿)"] else None,
            "Max Loss (฿)":  round(-STOP_LOSS*px*shares,0),
        })
    cash_remaining = FWD_CAPITAL - total_invested
    portfolio_df = pd.DataFrame(port_rows)

    # ── 22-day walk-forward tracker (uses last 22 biz days as historical proxy) ──
    # Simulate what would have happened if we ran this strategy 22 days ago
    tracker_rows = []
    positions = {}   # {ticker: (shares, entry_px)}
    fwd_cash  = FWD_CAPITAL
    for i, row in portfolio_df.iterrows():
        t   = row["Ticker"]
        n   = row["Name"]
        sh  = row["Shares"]
        px  = row["Entry Price(฿)"]
        if sh > 0 and t in stock_dict:
            # Get price 22 trading days ago as simulated entry
            _, df_t = stock_dict[t]
            entry_idx = max(0, len(df_t)-FWRD_DAYS-1)
            entry_px  = float(df_t["Close"].iloc[entry_idx])
            fwd_cash -= sh * entry_px * (1+TX_COST)
            positions[t] = {"name":n,"shares":sh,"avg_cost":entry_px}

    fwd_dates = None
    for t in positions:
        _, df_t = stock_dict[t]
        cand = df_t.index[-FWRD_DAYS:]
        if fwd_dates is None or len(cand)<len(fwd_dates):
            fwd_dates = cand

    if fwd_dates is not None:
        peak_val = FWD_CAPITAL
        for date in fwd_dates:
            day_pv = fwd_cash
            holding_details = []
            for t, h in positions.items():
                _, df_t = stock_dict[t]
                if date in df_t.index:
                    px = float(df_t.loc[date,"Close"])
                    day_pv += h["shares"]*px
                    unreal = (px - h["avg_cost"]) * h["shares"]
                    holding_details.append(f"{h['name']}:{unreal:+.0f}")
            if day_pv > peak_val: peak_val = day_pv
            dd = (day_pv-peak_val)/peak_val*100
            ret = (day_pv-FWD_CAPITAL)/FWD_CAPITAL*100
            tracker_rows.append({
                "Date":        date.date(),
                "Portfolio(฿)":round(day_pv,0),
                "Return (%)":  round(ret,2),
                "Drawdown(%)": round(dd,2),
                "Cash (฿)":    round(fwd_cash,0),
                "Holdings":    " | ".join(holding_details),
            })
    tracker_df = pd.DataFrame(tracker_rows)
    return signal_df, portfolio_df, tracker_df, cash_remaining

# ─── Excel report builder ─────────────────────────────────────────────────────
DARK  = "1F4E79"; MID   = "2E75B6"; LIGHT = "BDD7EE"
GRN   = "00B050"; RED   = "FF0000"; YLW   = "FFFF00"
LGRN  = "E2EFDA"; LRED  = "FCE4D6"; GRY   = "D9D9D9"

def build_excel(eq_df, bm_eq, trades_df, per_stock_df,
                fwd_signal_df, fwd_port_df, fwd_tracker_df, fwd_cash,
                strat_m, bm_m, run_date):
    wb = Workbook(); wb.remove(wb.active)

    # ══════════════════ SHEET 1: DASHBOARD ══════════════════
    ws = wb.create_sheet("Dashboard"); ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22; ws.column_dimensions["D"].width = 40

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Thai SET Strategy — Backtest & Forward Test Report"
    apply(ws["A1"], bold=True, sz=16, color="FFFFFF", bg=DARK, h="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    ws["A2"] = (f"Backtest: {BT_START.strftime('%d %b %Y')} → {BT_END.strftime('%d %b %Y')}  |  "
                f"Strategy: RSI({RSI_N}) + SMA{SMA_N} + MACD(12,26,9)  |  "
                f"Capital: ฿{CAPITAL:,}  |  Tx cost: {TX_COST*100:.2f}% per side")
    apply(ws["A2"], sz=9, color="595959", bg="DCE6F1", h="center")
    ws.row_dimensions[2].height = 18

    # Metrics table headers
    for c, txt, bg in [(1,"Metric",MID),(2,"Strategy",MID),(3,"Buy & Hold (Benchmark)",MID),(4,"Notes",MID)]:
        apply(ws.cell(4, c, txt), bold=True, sz=10, color="FFFFFF", bg=MID)
    ws.row_dimensions[4].height = 20

    metric_notes = {
        "Total Return (%)":      "Cumulative return over backtest period",
        "Annualised (%)":        "Return scaled to 1-year equivalent",
        "Volatility (%)":        "Annualised standard deviation of daily returns",
        "Sharpe Ratio":          "Return per unit of risk (risk-free = 1.5%)",
        "Max Drawdown (%)":      "Largest peak-to-trough decline",
        "Final Value (฿)":       f"Portfolio value starting from ฿{CAPITAL:,}",
        "P&L (฿)":               "Total profit or loss in Thai Baht",
    }

    keys = ["Total Return (%)","Annualised (%)","Volatility (%)","Sharpe Ratio",
            "Max Drawdown (%)","Final Value (฿)","P&L (฿)"]
    for i, k in enumerate(keys):
        r = 5 + i
        apply(ws.cell(r,1, k), sz=10, h="left")
        sv = strat_m.get(k,"—"); bv = bm_m.get(k,"—")

        # Strategy value
        sc = ws.cell(r, 2, sv)
        is_pct = "(%)" in k; is_money = "฿" in k
        fmt = "0.0%;(0.0%);-" if is_pct else ("#,##0;(#,##0);-" if is_money else "0.00")
        col = "000000"
        if isinstance(sv, (int,float)):
            if is_pct: sv_disp = sv/100; sc.value = sv_disp
            if sv < 0: col = "C00000"
            elif sv > 0 and k not in ["Volatility (%)"]: col = "375623"
        apply(sc, sz=10, color=col, fmt=fmt if not is_money else "#,##0;(#,##0);-")

        # Benchmark value
        bc = ws.cell(r, 3, bv)
        if isinstance(bv,(int,float)):
            if is_pct: bc.value = bv/100
            bc_col = "C00000" if bv<0 else "375623"
        else: bc_col = "000000"
        apply(bc, sz=10, color=bc_col, fmt=fmt if not is_money else "#,##0;(#,##0);-")
        apply(ws.cell(r,4, metric_notes.get(k,"")), sz=9, color="595959", h="left")
        ws.row_dimensions[r].height = 18

    # Alpha box
    r = 5 + len(keys) + 1
    ws.merge_cells(f"A{r}:D{r}")
    alpha = strat_m.get("Total Return (%)",0) - bm_m.get("Total Return (%)",0)
    ws[f"A{r}"] = (f"⚡  Strategy Alpha vs Benchmark: {alpha:+.2f}%  "
                   f"({'Outperformed' if alpha>0 else 'Underperformed'} by {abs(alpha):.2f}%)")
    apply(ws[f"A{r}"], bold=True, sz=11,
          color="FFFFFF" if alpha>=0 else "FFFFFF",
          bg=GRN if alpha>=0 else "C00000", h="center")
    ws.row_dimensions[r].height = 26

    # Disclaimer
    r2 = r+2
    ws.merge_cells(f"A{r2}:D{r2}")
    ws[f"A{r2}"] = ("⚠  DISCLAIMER: This backtest uses historical data. Past performance does not guarantee "
                    "future results. Signals are for educational purposes only — not financial advice.")
    apply(ws[f"A{r2}"], sz=9, color="C00000", bg="FCE4D6", h="left", wrap=True)
    ws.row_dimensions[r2].height = 28

    # ══════════════════ SHEET 2: EQUITY CURVE ══════════════════
    ws2 = wb.create_sheet("Equity Curve"); ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Portfolio Value vs Benchmark (฿1,000,000 start)"
    apply(ws2["A1"], bold=True, sz=14, color="FFFFFF", bg=DARK, h="center")
    ws2.row_dimensions[1].height = 30

    for c, h, w in [(1,"Date",13),(2,"Strategy (฿)",15),(3,"Return (%)",13),
                     (4,"Benchmark (฿)",16),(5,"BM Return (%)",14),(6,"Daily P&L (฿)",14)]:
        apply(ws2.cell(2,c,h), bold=True, sz=9, color="FFFFFF", bg=MID)
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[2].height = 20

    eq_aligned = eq_df.reindex(bm_eq.index, method="ffill").fillna(method="bfill") if bm_eq is not None else eq_df
    dates_list  = list(eq_aligned.index if bm_eq is not None else eq_df.index)
    start_port  = float(eq_df["Portfolio"].iloc[0])
    start_bm    = float(bm_eq["Benchmark"].iloc[0]) if bm_eq is not None else None

    for i, d in enumerate(dates_list):
        r = i + 3
        sv = float(eq_aligned.loc[d,"Portfolio"]) if d in eq_aligned.index else None
        bv = float(bm_eq.loc[d,"Benchmark"]) if bm_eq is not None and d in bm_eq.index else None
        ws2.cell(r,1,d.date()).number_format = "DD-MMM-YY"
        ws2.cell(r,1).alignment = Al(h="center"); ws2.cell(r,1).font=F(sz=9); ws2.cell(r,1).border=Brd()
        if sv is not None:
            apply(ws2.cell(r,2,sv), sz=9, fmt="#,##0;(#,##0);-")
            ret_pct = (sv/start_port-1)
            apply(ws2.cell(r,3,ret_pct), sz=9, fmt="0.0%;(0.0%);-",
                  color="375623" if ret_pct>=0 else "C00000")
        if bv is not None:
            apply(ws2.cell(r,4,bv), sz=9, fmt="#,##0;(#,##0);-")
            bm_ret = (bv/start_bm-1)
            apply(ws2.cell(r,5,bm_ret), sz=9, fmt="0.0%;(0.0%);-",
                  color="375623" if bm_ret>=0 else "C00000")
        if sv is not None and i>0:
            ws2.cell(r,6, f"=B{r}-B{r-1}").number_format="#,##0;(#,##0);-"
            ws2.cell(r,6).font=F(sz=9); ws2.cell(r,6).border=Brd(); ws2.cell(r,6).alignment=Al()

    # Equity chart
    last_data_row = len(dates_list) + 2
    chart = LineChart(); chart.title="Portfolio vs Benchmark"; chart.style=2
    chart.height=14; chart.width=26; chart.grouping="standard"; chart.smooth=True
    chart.y_axis.title="Value (฿)"; chart.x_axis.title="Date"
    port_ref = Reference(ws2,min_col=2,min_row=2,max_row=last_data_row)
    chart.add_data(port_ref,titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill=MID; chart.series[0].graphicalProperties.line.width=18000
    if bm_eq is not None:
        bm_ref = Reference(ws2,min_col=4,min_row=2,max_row=last_data_row)
        chart.add_data(bm_ref,titles_from_data=True)
        chart.series[1].graphicalProperties.line.solidFill="ED7D31"; chart.series[1].graphicalProperties.line.width=12000
    ws2.add_chart(chart,"H3")

    # ══════════════════ SHEET 3: STOCK RESULTS ══════════════════
    ws3 = wb.create_sheet("Stock Results"); ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:L1")
    ws3["A1"] = "Individual Stock Backtest Results"
    apply(ws3["A1"], bold=True, sz=14, color="FFFFFF", bg=DARK, h="center")
    ws3.row_dimensions[1].height = 30

    ps_cols = ["Ticker","Name","Buy-Hold (%)","# Trades","Wins","Losses",
               "Win Rate (%)","Current Signal","Score","Price (฿)","RSI","SMA50 (฿)"]
    ps_widths= [12,18,14,10,8,8,12,14,8,12,8,12]
    for c,(h,w) in enumerate(zip(ps_cols,ps_widths),1):
        apply(ws3.cell(2,c,h), bold=True, sz=9, color="FFFFFF", bg=MID)
        ws3.column_dimensions[get_column_letter(c)].width=w
    ws3.row_dimensions[2].height=20

    if not per_stock_df.empty:
        for i,row in per_stock_df.iterrows():
            r=i+3
            for c,col in enumerate(ps_cols,1):
                v=row.get(col,""); cell=ws3.cell(r,c,v)
                fmt=None
                if col in ["Buy-Hold (%)","Win Rate (%)"]:
                    fmt="0.0%;(0.0%);-"
                    if pd.notna(v) and isinstance(v,(int,float)): cell.value=v/100
                apply(cell,sz=9,h="center",fmt=fmt)
                if col=="Current Signal":
                    if v in ("STRONG BUY","BUY"):
                        cell.fill=Fill("E2EFDA"); cell.font=F(sz=9,bold=True,color="375623")
                    elif v in ("STRONG SELL","SELL"):
                        cell.fill=Fill("FCE4D6"); cell.font=F(sz=9,bold=True,color="C00000")
                    else:
                        cell.fill=Fill(GRY)
            ws3.row_dimensions[r].height=16

    # ══════════════════ SHEET 4: TRADE LOG ══════════════════
    ws4 = wb.create_sheet("Trade Log"); ws4.sheet_view.showGridLines=False
    ws4.merge_cells("A1:H1")
    ws4["A1"]="Full Trade Log — All Transactions"
    apply(ws4["A1"],bold=True,sz=14,color="FFFFFF",bg=DARK,h="center")
    ws4.row_dimensions[1].height=30

    tl_cols=["Date","Ticker","Name","Action","Price (฿)","Shares","Value (฿)","Tx Cost (฿)"]
    tl_widths=[13,10,18,8,12,10,14,13]
    for c,(h,w) in enumerate(zip(tl_cols,tl_widths),1):
        apply(ws4.cell(2,c,h),bold=True,sz=9,color="FFFFFF",bg=MID)
        ws4.column_dimensions[get_column_letter(c)].width=w
    ws4.row_dimensions[2].height=20

    if not trades_df.empty:
        for i,row in trades_df.iterrows():
            r=i+3; action=row.get("Action","")
            data=[(1,row.get("Date",""),"DD-MMM-YY"),
                  (2,row.get("Ticker",""),None),(3,row.get("Name",""),None),
                  (4,action,None),(5,row.get("Price",0),"#,##0.00"),
                  (6,row.get("Shares",0),"#,##0"),
                  (7,row.get("Value",0),"#,##0;(#,##0);-"),
                  (8,row.get("TxCost",0),"#,##0;(#,##0);-")]
            for c,v,fmt in data:
                cell=ws4.cell(r,c,v); apply(cell,sz=9,fmt=fmt)
                if c==4:
                    cell.fill=Fill("E2EFDA") if action=="BUY" else Fill("FCE4D6")
                    cell.font=F(sz=9,bold=True,color="375623" if action=="BUY" else "C00000")
            ws4.row_dimensions[r].height=16

        # Summary totals
        tr = len(trades_df)+4
        ws4.cell(tr,3,"TOTAL TRANSACTIONS").font=F(bold=True,sz=10)
        ws4.cell(tr,5,f"=COUNTA(D3:D{tr-1})").font=F(sz=10)
        ws4.cell(tr,7,f'=SUMIF(D3:D{tr-1},"SELL",G3:G{tr-1})-SUMIF(D3:D{tr-1},"BUY",G3:G{tr-1})')
        ws4.cell(tr,7).number_format="#,##0;(#,##0);-"; ws4.cell(tr,7).font=F(bold=True,sz=10)
        ws4.cell(tr,8,f"=SUM(H3:H{tr-1})").number_format="#,##0"; ws4.cell(tr,8).font=F(bold=True,sz=10)

    # ══════════════════ SHEET 5: FORWARD PORTFOLIO (฿300k plan) ══════════════════
    ws5 = wb.create_sheet("Forward Portfolio"); ws5.sheet_view.showGridLines=False

    # ── Header ──
    ws5.merge_cells("A1:L1")
    ws5["A1"] = (f"Forward Portfolio — ฿{FWD_CAPITAL:,} Starting Capital  "
                 f"|  Top-{FWD_MAX_POS} BUY Positions  |  {TODAY.strftime('%d %b %Y')}")
    apply(ws5["A1"],bold=True,sz=13,color="FFFFFF",bg=DARK,h="center")
    ws5.row_dimensions[1].height=32

    # Capital summary box
    ws5.merge_cells("A2:D2"); ws5.merge_cells("E2:H2"); ws5.merge_cells("I2:L2")
    total_inv = fwd_port_df["Invested (฿)"].sum() if not fwd_port_df.empty else 0
    apply(ws5.cell(2,1, f"💼 Total Capital: ฿{FWD_CAPITAL:,}"),
          bold=True, sz=10, color="FFFFFF", bg=DARK, h="center")
    apply(ws5.cell(2,5, f"📊 Invested: ฿{total_inv:,.0f}  ({total_inv/FWD_CAPITAL*100:.1f}%)"),
          bold=True, sz=10, color="FFFFFF", bg="375623", h="center")
    apply(ws5.cell(2,9, f"💰 Cash Reserve: ฿{fwd_cash:,.0f}  ({fwd_cash/FWD_CAPITAL*100:.1f}%)"),
          bold=True, sz=10, color="FFFFFF", bg=MID, h="center")
    ws5.row_dimensions[2].height=24

    ws5.merge_cells("A3:L3")
    ws5["A3"] = (f"⚠  BUY targets = +3%  |  Stop-loss = -{STOP_LOSS*100:.0f}%  "
                 f"|  Lot size = {FWD_LOT} shares  |  Commission = {TX_COST*100:.2f}% per side  "
                 f"|  For educational purposes only — not financial advice")
    apply(ws5["A3"],sz=9,color="C00000",bg="FCE4D6",h="left",wrap=True)
    ws5.row_dimensions[3].height=20

    # Portfolio table
    fp_cols   = ["Ticker","Name","Signal","Score","Entry Price(฿)","Shares",
                 "Invested (฿)","Weight (%)","1M Target (฿)","Stop-Loss (฿)",
                 "Target P&L (฿)","Max Loss (฿)"]
    fp_widths = [11,16,14,7,14,8,14,10,14,14,14,12]
    for c,(h,w) in enumerate(zip(fp_cols,fp_widths),1):
        bg_c = GRN if h in ("1M Target (฿)","Target P&L (฿)") else (
               "C00000" if h=="Max Loss (฿)" else MID)
        apply(ws5.cell(4,c,h),bold=True,sz=9,color="FFFFFF",bg=bg_c)
        ws5.column_dimensions[get_column_letter(c)].width=w
    ws5.row_dimensions[4].height=20

    if not fwd_port_df.empty:
        for i,row in fwd_port_df.iterrows():
            r = i+5; sig=row.get("Signal","")
            for c,col in enumerate(fp_cols,1):
                v = row.get(col,""); cell=ws5.cell(r,c,v)
                fmt=None
                if col in ["Entry Price(฿)","1M Target (฿)","Stop-Loss (฿)"]:
                    fmt="#,##0.00"
                elif col in ["Invested (฿)","Target P&L (฿)","Max Loss (฿)"]:
                    fmt="#,##0;(#,##0);-"
                elif col=="Weight (%)":
                    fmt="0.0\"%\""
                elif col=="Shares": fmt="#,##0"
                apply(cell,sz=9,h="center" if col not in ("Name","Ticker") else "left",fmt=fmt)
                if col=="Signal":
                    if "BUY" in str(sig):
                        cell.fill=Fill("E2EFDA"); cell.font=F(sz=9,bold=True,color="375623")
                if col=="Target P&L (฿)" and isinstance(v,(int,float)) and v>0:
                    cell.fill=Fill("E2EFDA"); cell.font=F(sz=9,color="375623")
                if col=="Max Loss (฿)" and isinstance(v,(int,float)) and v<0:
                    cell.fill=Fill("FCE4D6"); cell.font=F(sz=9,color="C00000")
            ws5.row_dimensions[r].height=17

        # Totals row
        tr = len(fwd_port_df)+5
        ws5.cell(tr,1,"TOTAL").font=F(bold=True,sz=10)
        for c,col in enumerate(fp_cols,1):
            if col in ["Invested (฿)","Target P&L (฿)","Max Loss (฿)"]:
                cell=ws5.cell(tr,c,f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{tr-1})")
                cell.number_format="#,##0;(#,##0);-"; cell.font=F(bold=True,sz=10)
                cell.border=Brd(); cell.alignment=Al()

    # ══════════════════ SHEET 6: PORTFOLIO TRACKER (22-day simulation) ══════════
    ws6 = wb.create_sheet("Portfolio Tracker"); ws6.sheet_view.showGridLines=False

    ws6.merge_cells("A1:F1")
    ws6["A1"] = (f"22-Day Walk-Forward Simulation  |  Capital: ฿{FWD_CAPITAL:,}  "
                 f"|  (Uses last 22 trading days of historical data as proxy)")
    apply(ws6["A1"],bold=True,sz=12,color="FFFFFF",bg=DARK,h="center")
    ws6.row_dimensions[1].height=30

    ws6.merge_cells("A2:F2")
    ws6["A2"] = ("ℹ  This simulation replays the model signals over the past 22 trading days. "
                 "It shows what the portfolio would have looked like — "
                 "to help you understand day-to-day performance patterns.")
    apply(ws6["A2"],sz=9,color="595959",bg=LIGHT,h="left",wrap=True)
    ws6.row_dimensions[2].height=24

    tr_cols  = ["Date","Portfolio(฿)","Return (%)","Drawdown(%)","Cash (฿)","Holdings P&L detail"]
    tr_widths= [13,15,12,12,14,60]
    for c,(h,w) in enumerate(zip(tr_cols,tr_widths),1):
        apply(ws6.cell(3,c,h),bold=True,sz=9,color="FFFFFF",bg=MID)
        ws6.column_dimensions[get_column_letter(c)].width=w
    ws6.row_dimensions[3].height=20

    if not fwd_tracker_df.empty:
        for i,row in fwd_tracker_df.iterrows():
            r=i+4; ret=row.get("Return (%)",0); dd=row.get("Drawdown(%)",0)
            d=ws6.cell(r,1,row.get("Date","")); d.number_format="DD-MMM-YY"
            d.alignment=Al(h="center"); d.font=F(sz=9); d.border=Brd()
            pv=ws6.cell(r,2,row.get("Portfolio(฿)",0)); apply(pv,sz=9,fmt="#,##0")
            rc=ws6.cell(r,3,ret/100); apply(rc,sz=9,fmt="0.00%;(0.00%);-",
                color="375623" if ret>=0 else "C00000")
            dc=ws6.cell(r,4,dd/100); apply(dc,sz=9,fmt="0.00%;(0.00%);-",
                color="C00000" if dd<-3 else ("FF6600" if dd<0 else "375623"))
            apply(ws6.cell(r,5,row.get("Cash (฿)",0)),sz=9,fmt="#,##0")
            hc=ws6.cell(r,6,str(row.get("Holdings",""))); hc.font=F(sz=8,color="595959")
            hc.alignment=Alignment(horizontal="left",vertical="center")
            hc.border=Brd()
            ws6.row_dimensions[r].height=16

        # Mini equity chart for tracker
        last_tr_row = len(fwd_tracker_df)+3
        chart2=LineChart(); chart2.title="Forward Portfolio Value (22-day sim)"
        chart2.style=2; chart2.height=12; chart2.width=20
        chart2.y_axis.title="Value (฿)"; chart2.x_axis.title="Day"
        p_ref=Reference(ws6,min_col=2,min_row=3,max_row=last_tr_row)
        chart2.add_data(p_ref,titles_from_data=True)
        chart2.series[0].graphicalProperties.line.solidFill=GRN
        chart2.series[0].graphicalProperties.line.width=18000
        ws6.add_chart(chart2,"H4")

    # ══════════════════ SHEET 7: ALL SIGNALS ══════════════════
    ws7 = wb.create_sheet("All Signals"); ws7.sheet_view.showGridLines=False
    ws7.merge_cells("A1:L1")
    ws7["A1"]="All Stock Signals (current as of run date)"
    apply(ws7["A1"],bold=True,sz=13,color="FFFFFF",bg=DARK,h="center")
    ws7.row_dimensions[1].height=28

    fw_cols=["Ticker","Name","Action","Signal","Score",
             "Price (฿)","SMA50 (฿)","RSI","MACD",
             "1M Target (฿)","Stop-Loss (฿)","Upside (%)"]
    fw_widths=[11,18,12,14,7,12,12,8,10,13,13,10]
    fw_cols_actual=["Ticker","Name","Action","Signal","Score",
                    "Current Price (฿)","SMA50 (฿)","RSI","MACD",
                    "1-Month Target (฿)","Stop-Loss (฿)","Upside (%)"]
    for c,(h,w) in enumerate(zip(fw_cols,fw_widths),1):
        apply(ws7.cell(2,c,h),bold=True,sz=9,color="FFFFFF",
              bg=GRN if h in ("Action","1M Target (฿)","Upside (%)") else MID)
        ws7.column_dimensions[get_column_letter(c)].width=w
    ws7.row_dimensions[2].height=20

    if not fwd_signal_df.empty:
        for i,row in fwd_signal_df.iterrows():
            r=i+3; action=row.get("Action","HOLD")
            for c,col in enumerate(fw_cols_actual,1):
                v=row.get(col,""); cell=ws7.cell(r,c,v)
                fmt=None
                if col in ["Current Price (฿)","SMA50 (฿)","1-Month Target (฿)","Stop-Loss (฿)"]:
                    fmt="#,##0.00;(#,##0.00);-"
                elif col=="Upside (%)":
                    fmt="0.0%;(0.0%);-"
                    if pd.notna(v) and isinstance(v,(int,float)): cell.value=v/100
                elif col=="MACD": fmt="0.0000;(0.0000);-"
                apply(cell,sz=9,fmt=fmt)
                if col=="Action":
                    if action=="BUY":
                        cell.fill=Fill("E2EFDA"); cell.font=F(sz=9,bold=True,color=GRN)
                    elif action=="AVOID/EXIT":
                        cell.fill=Fill("FCE4D6"); cell.font=F(sz=9,bold=True,color="C00000")
                    else:
                        cell.fill=Fill(GRY)
                if col=="Signal":
                    if "BUY" in str(v): cell.fill=Fill("E2EFDA"); cell.font=F(sz=9,color="375623")
                    elif "SELL" in str(v): cell.fill=Fill("FCE4D6"); cell.font=F(sz=9,color="C00000")
            ws7.row_dimensions[r].height=16

    wb.save(OUTPUT_PATH)
    print(f"\n✅ Report saved: {OUTPUT_PATH}")
    return OUTPUT_PATH

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("="*60)
    print("Thai SET Strategy — Backtest & Forward Test")
    print(f"Run: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Period: {BT_START.strftime('%d %b %Y')} → {BT_END.strftime('%d %b %Y')}")
    print(f"Stocks: {len(INSTRUMENTS)}  |  Capital: ฿{CAPITAL:,}")
    print("="*60)

    all_instruments = list(dict.fromkeys([BENCHMARK] + INSTRUMENTS))  # benchmark first

    print(f"\nFetching {len(all_instruments)} instruments ({FETCH_PERIOD})...\n")
    raw = {}
    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(fetch, n, t): (n,t) for n,t in all_instruments}
        for f in as_completed(futures):
            n,t,df = f.result()
            status = f"✓ {len(df)} rows" if df is not None else "✗ failed"
            print(f"  {n:15s} ({t:12s}) {status}")
            if df is not None: raw[(n,t)] = df

    if len(raw) < 2:
        print("\nERROR: Insufficient data fetched. Check internet connection."); sys.exit(1)

    # Separate benchmark
    bm_df = raw.get(BENCHMARK)
    stock_raw = {t:(n,df) for (n,t),df in raw.items() if (n,t)!=BENCHMARK}

    # Build signal DataFrames
    print("\nCalculating indicators & signals...")
    stock_dict = {}
    for t,(n,df) in stock_raw.items():
        sig_df = add_signals(df["Close"], df["Open"] if "Open" in df.columns else df["Close"])
        stock_dict[t] = (n, sig_df)

    # Run backtest
    print(f"\nRunning portfolio backtest ({BT_START.date()} → {BT_END.date()})...")
    eq_df, trades_df, per_stock_df = backtest(stock_dict)

    # Benchmark equity curve (normalised to same capital)
    bm_eq = None
    if bm_df is not None:
        bm_bt = bm_df[(bm_df.index>=BT_START)&(bm_df.index<=BT_END)]["Close"].dropna()
        if len(bm_bt)>1:
            bm_norm = bm_bt / bm_bt.iloc[0] * CAPITAL
            bm_eq   = pd.DataFrame({"Benchmark": bm_norm.round(0)})

    # Performance metrics
    strat_m = perf(eq_df["Portfolio"].astype(float), "Strategy") if eq_df is not None else {}
    bm_m    = perf(bm_eq["Benchmark"].astype(float), "Benchmark") if bm_eq is not None else {}

    print("\n── Performance Summary ──")
    for k in ["Total Return (%)","Annualised (%)","Sharpe Ratio","Max Drawdown (%)"]:
        sv = strat_m.get(k,"—"); bv = bm_m.get(k,"—")
        print(f"  {k:26s}  Strategy: {str(sv):>8s}    Benchmark: {str(bv):>8s}")

    # Forward test
    print("\nRunning forward test (current signals)...")
    fwd_signal_df, fwd_port_df, fwd_tracker_df, fwd_cash = forward_test(stock_dict)
    buys  = fwd_signal_df[fwd_signal_df["Action"]=="BUY"]
    sells = fwd_signal_df[fwd_signal_df["Action"]=="AVOID/EXIT"]
    print(f"  BUY signals:    {len(buys)}  |  SELL/AVOID: {len(sells)}")
    print(f"  Portfolio plan: {len(fwd_port_df)} positions  "
          f"|  Invested: ฿{fwd_port_df['Invested (฿)'].sum():,.0f}  "
          f"|  Cash reserve: ฿{fwd_cash:,.0f}")

    # Build Excel report
    print("\nBuilding Excel report...")
    build_excel(eq_df, bm_eq, trades_df, per_stock_df,
                fwd_signal_df, fwd_port_df, fwd_tracker_df, fwd_cash,
                strat_m, bm_m, TODAY)

    print(f"\nDone! Open: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
