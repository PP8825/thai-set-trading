#!/usr/bin/env python3
"""
Generate set_backtest_report.xlsx from set_backtest_results.json
"""
import json, datetime, math
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import DataPoint

SRC = Path(__file__).parent / "set_backtest_results.json"
OUT = Path(__file__).parent / "set_backtest_report.xlsx"

with open(SRC) as f:
    data = json.load(f)

perf   = data["performance"]
trades = data["trades"]
equity = data["equity_curve"]
meta   = data["meta"]

# ── Colour palette ────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_GREEN      = "217346"
C_RED        = "C00000"
C_AMBER      = "ED7D31"
C_GREY_HDR   = "404040"
C_GREY_ROW   = "F2F2F2"
C_WHITE      = "FFFFFF"
C_GOLD       = "FFC000"
C_BULL       = "C6EFCE"
C_BEAR       = "FFCCCC"

# ── Style helpers ─────────────────────────────────────────────────────────────
def hfont(bold=True, size=11, colour=C_WHITE, name="Arial"):
    return Font(bold=bold, size=size, color=colour, name=name)

def bfont(bold=False, size=10, colour="000000", name="Arial"):
    return Font(bold=bold, size=size, color=colour, name=name)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def right():
    return Alignment(horizontal="right", vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    return Border(bottom=Side(style="medium", color="000000"))

def style_header_row(ws, row, cols, bg=C_DARK_BLUE, fg=C_WHITE, sz=10):
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, size=sz, color=fg, name="Arial")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    bg = C_GREY_ROW if alt else C_WHITE
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        c.font = bfont()
        c.fill = fill(bg)
        c.border = thin_border()

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def pct_fmt(v, pos=True):
    sign = "+" if (pos and v >= 0) else ""
    return f"{sign}{v:.1f}%"

# ── Build DataFrames ──────────────────────────────────────────────────────────
df_trades = pd.DataFrame(trades)
df_eq     = pd.DataFrame(equity)
df_eq["date"] = pd.to_datetime(df_eq["date"])
df_eq["month"] = df_eq["date"].dt.to_period("M")

sells = df_trades[df_trades["action"] == "SELL"].copy()
sells = sells[sells["reason"] != "End of backtest"].copy()
buys  = df_trades[df_trades["action"] == "BUY"].copy()

# ── Monthly returns ───────────────────────────────────────────────────────────
df_eq_sorted = df_eq.sort_values("date")
monthly = (df_eq_sorted.groupby("month")["value"]
           .agg(["first", "last"])
           .reset_index())
monthly.columns = ["month", "start_val", "end_val"]
monthly["return_pct"] = (monthly["end_val"] - monthly["start_val"]) / monthly["start_val"] * 100
monthly["year"]  = monthly["month"].dt.year
monthly["month_num"] = monthly["month"].dt.month

# ── Per-ticker stats ──────────────────────────────────────────────────────────
ticker_stats = []
for tk in sells["ticker"].unique():
    ts = sells[sells["ticker"] == tk]
    wins   = ts[ts["pnl"] > 0]
    losses = ts[ts["pnl"] <= 0]
    ticker_stats.append({
        "Ticker":        tk,
        "Name":          ts["name"].iloc[0],
        "Trades":        len(ts),
        "Wins":          len(wins),
        "Win Rate":      len(wins)/len(ts)*100 if len(ts) else 0,
        "Total PnL":     ts["pnl"].sum(),
        "Avg Win":       wins["pnl"].mean() if len(wins) else 0,
        "Avg Loss":      losses["pnl"].mean() if len(losses) else 0,
        "Best Trade":    ts["pnl"].max(),
        "Worst Trade":   ts["pnl"].min(),
        "Avg Hold Days": ts["hold_days"].mean() if len(ts) else 0,
    })
df_tk = pd.DataFrame(ticker_stats).sort_values("Total PnL", ascending=False)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False

# Title banner
ws.merge_cells("A1:H1")
ws["A1"] = "SET Strategy Backtest Report"
ws["A1"].font = Font(bold=True, size=18, color=C_WHITE, name="Arial")
ws["A1"].fill = fill(C_DARK_BLUE)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:H2")
ws["A2"] = (f"Period: {equity[0]['date']}  →  {equity[-1]['date']}   |   "
            f"Capital: ฿{meta['capital']:,.0f}   |   "
            f"Run date: {meta['run_date']}   |   "
            f"Regime filter: {'ON' if meta['regime'] else 'OFF'}")
ws["A2"].font = Font(size=10, color=C_LIGHT_BLUE, italic=True, name="Arial")
ws["A2"].fill = fill(C_DARK_BLUE)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 20

ws.row_dimensions[3].height = 10

# ── Performance metrics grid ──────────────────────────────────────────────────
metrics = [
    ("RETURNS", [
        ("Initial Capital",   f"฿{meta['capital']:,.0f}",          False),
        ("Final Value",        f"฿{perf['final_value']:,.0f}",      False),
        ("Total Return",       pct_fmt(perf['total_return']),        True),
        ("Annual Return",      pct_fmt(perf['annual_return']),       True),
        ("SET B&H Return",     pct_fmt(perf['set_bah_return'] or 0), True),
        ("Alpha vs SET",       pct_fmt(perf['alpha']),               True),
        ("Dividend Income",    f"฿{perf.get('dividend_income',0):,.0f}", False),
    ]),
    ("RISK", [
        ("Sharpe Ratio",       f"{perf['sharpe']:.2f}",            False),
        ("Max Drawdown",       pct_fmt(perf['max_drawdown'],False), True),
        ("Profit Factor",      f"{perf['profit_factor']:.2f}x",    False),
        ("Avg Hold Days",      f"{perf['avg_hold_days']:.0f} days", False),
    ]),
    ("TRADES", [
        ("Total Trades",       f"{perf['n_trades']}",              False),
        ("Win Rate",           f"{perf['win_rate']:.1f}%",         False),
        ("Winning Trades",     f"{len(sells[sells['pnl']>0])}",    False),
        ("Losing Trades",      f"{len(sells[sells['pnl']<=0])}",   False),
        ("Avg Win  (฿)",       f"฿{sells[sells['pnl']>0]['pnl'].mean():,.0f}" if len(sells[sells['pnl']>0]) else "–", False),
        ("Avg Loss (฿)",       f"฿{sells[sells['pnl']<=0]['pnl'].mean():,.0f}" if len(sells[sells['pnl']<=0]) else "–", False),
        ("Best Trade (฿)",     f"฿{sells['pnl'].max():,.0f}",      False),
        ("Worst Trade (฿)",    f"฿{sells['pnl'].min():,.0f}",      False),
    ]),
]

start_row = 4
col_offsets = [1, 4, 7]  # A, D, G

for (section, rows), col_off in zip(metrics, col_offsets):
    # Section header
    ws.merge_cells(start_row=start_row, start_column=col_off,
                   end_row=start_row,   end_column=col_off+1)
    c = ws.cell(row=start_row, column=col_off, value=section)
    c.font  = Font(bold=True, size=10, color=C_WHITE, name="Arial")
    c.fill  = fill(C_MID_BLUE)
    c.alignment = center()
    ws.cell(row=start_row, column=col_off+1).fill = fill(C_MID_BLUE)

    for i, (label, value, is_pct) in enumerate(rows):
        r = start_row + 1 + i
        lc = ws.cell(row=r, column=col_off, value=label)
        vc = ws.cell(row=r, column=col_off+1, value=value)
        bg = C_GREY_ROW if i % 2 == 0 else C_WHITE
        lc.fill = fill(bg); vc.fill = fill(bg)
        lc.font = bfont(); vc.font = bfont(bold=True)
        lc.alignment = left(); vc.alignment = right()
        lc.border = thin_border(); vc.border = thin_border()

        # Colour-code pct values
        if is_pct and value and value != "–":
            num = float(value.replace("+","").replace("%",""))
            vc.font = Font(bold=True, size=10, name="Arial",
                          color=C_GREEN if num >= 0 else C_RED)

# ── Column widths ─────────────────────────────────────────────────────────────
set_col_widths(ws, [18, 14, 2, 18, 14, 2, 18, 14])

# ── Mini equity chart on Summary ──────────────────────────────────────────────
# We'll add a proper chart on the equity sheet; put a note here
chart_row = start_row + max(len(m[1]) for m in metrics) + 3

ws.merge_cells(start_row=chart_row, start_column=1, end_row=chart_row, end_column=8)
ws.cell(row=chart_row, column=1,
        value="📈  See 'Equity Curve' sheet for portfolio growth chart").font = \
    Font(italic=True, size=10, color=C_MID_BLUE, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Equity Curve
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Equity Curve")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
ws2["A1"] = "Daily Portfolio Equity Curve"
ws2["A1"].font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
ws2["A1"].fill = fill(C_DARK_BLUE)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 28

headers = ["Date", "Portfolio Value (฿)", "Cash (฿)", "# Positions", "Regime",
           "Daily Change (฿)", "Daily Chg %"]
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font  = Font(bold=True, size=9, color=C_WHITE, name="Arial")
    c.fill  = fill(C_GREY_HDR)
    c.alignment = center()
    c.border = thin_border()

eq_sorted = sorted(equity, key=lambda x: x["date"])
for i, row in enumerate(eq_sorted):
    r = i + 3
    alt = i % 2 == 0
    regime = row.get("regime", "BULL")
    bg = C_BULL if regime == "BULL" else C_BEAR

    ws2.cell(r, 1, row["date"]).number_format = "YYYY-MM-DD"
    ws2.cell(r, 2, row["value"]).number_format = "#,##0"
    ws2.cell(r, 3, row["cash"]).number_format  = "#,##0"
    ws2.cell(r, 4, row["n"])
    ws2.cell(r, 5, regime)
    # Daily change formula (skip row 3 which has no prior)
    if r > 3:
        ws2.cell(r, 6, f"=B{r}-B{r-1}").number_format = "#,##0;(#,##0);-"
        ws2.cell(r, 7, f"=IF(B{r-1}<>0,(B{r}-B{r-1})/B{r-1},0)").number_format = "0.00%"
    else:
        ws2.cell(r, 6, 0).number_format = "#,##0;(#,##0);-"
        ws2.cell(r, 7, 0).number_format = "0.00%"

    for col in range(1, 8):
        c = ws2.cell(r, col)
        c.font   = Font(size=9, name="Arial")
        c.fill   = fill(bg if col == 5 else (C_GREY_ROW if alt else C_WHITE))
        c.border = thin_border()
        c.alignment = center() if col in (1, 4, 5) else right()

# Chart
n_rows = len(eq_sorted)
chart = LineChart()
chart.title = "Portfolio Value vs Initial Capital"
chart.style = 10
chart.width = 26
chart.height = 14
chart.y_axis.title = "Portfolio Value (฿)"
chart.x_axis.title = "Trading Days"
chart.y_axis.numFmt = "#,##0"

data_ref = Reference(ws2, min_col=2, min_row=2, max_row=n_rows+2)
chart.add_data(data_ref, titles_from_data=True)
chart.series[0].graphicalProperties.line.solidFill = C_MID_BLUE
chart.series[0].graphicalProperties.line.width = 18000

# Add initial capital reference line (flat)
from openpyxl.chart.series import Series
ws2["I1"] = "Initial Capital"
for i in range(n_rows):
    ws2.cell(i+3, 9, meta["capital"])
ref2 = Reference(ws2, min_col=9, min_row=2, max_row=n_rows+2)
chart.add_data(ref2, titles_from_data=True)
chart.series[1].graphicalProperties.line.solidFill = C_AMBER
chart.series[1].graphicalProperties.line.width = 12000
chart.series[1].graphicalProperties.line.dashDot = "dash"

ws2.add_chart(chart, "K2")
set_col_widths(ws2, [12, 16, 14, 12, 8, 16, 12, 2, 16])

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Monthly Returns
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Monthly Returns")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:N1")
ws3["A1"] = "Monthly Returns Heatmap"
ws3["A1"].font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
ws3["A1"].fill = fill(C_DARK_BLUE)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 28

month_names = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec","FULL YEAR"]

# Header row
ws3.cell(2, 1, "Year").font = Font(bold=True, size=10, color=C_WHITE, name="Arial")
ws3.cell(2, 1).fill = fill(C_GREY_HDR)
ws3.cell(2, 1).alignment = center()
for col, mn in enumerate(month_names, 2):
    c = ws3.cell(2, col, mn)
    c.font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
    c.fill = fill(C_GREY_HDR)
    c.alignment = center()

years = sorted(monthly["year"].unique())
monthly_pivot = {}
for _, row in monthly.iterrows():
    monthly_pivot[(row["year"], row["month_num"])] = row["return_pct"]

for yi, year in enumerate(years):
    r = yi + 3
    ws3.cell(r, 1, year).font = Font(bold=True, size=10, name="Arial")
    ws3.cell(r, 1).fill = fill(C_LIGHT_BLUE)
    ws3.cell(r, 1).alignment = center()

    year_total = 0.0
    for mi in range(1, 13):
        val = monthly_pivot.get((year, mi))
        c = ws3.cell(r, mi+1)
        if val is not None:
            c.value = val / 100
            c.number_format = "0.0%"
            c.font = Font(bold=True, size=9, name="Arial",
                         color=C_WHITE if abs(val) > 1 else "000000")
            if val >= 3:   c.fill = fill("00B050")
            elif val >= 1: c.fill = fill("92D050")
            elif val >= 0: c.fill = fill("C6EFCE")
            elif val >= -1:c.fill = fill("FFCCCC")
            elif val >= -3:c.fill = fill("FF7676")
            else:          c.fill = fill(C_RED)
            year_total += val
        else:
            c.value = "–"
            c.fill = fill(C_GREY_ROW)
            c.font = Font(size=9, color="999999", name="Arial")
        c.alignment = center()
        c.border = thin_border()

    # Full year cell
    fc = ws3.cell(r, 14, year_total/100)
    fc.number_format = "0.0%"
    fc.font = Font(bold=True, size=9, name="Arial",
                  color=C_WHITE if abs(year_total) > 2 else "000000")
    fc.fill = fill("00B050" if year_total >= 0 else C_RED)
    fc.alignment = center()
    fc.border = thin_border()

set_col_widths(ws3, [8] + [8]*13)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Trade Log
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Trade Log")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:J1")
ws4["A1"] = "Complete Trade Log"
ws4["A1"].font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
ws4["A1"].fill = fill(C_DARK_BLUE)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 28

log_headers = ["Date", "Action", "Ticker", "Name", "Shares",
               "Price (฿)", "Avg Cost (฿)", "P&L (฿)", "Return %", "Reason", "Hold Days"]
for col, h in enumerate(log_headers, 1):
    c = ws4.cell(2, col, h)
    c.font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
    c.fill = fill(C_GREY_HDR)
    c.alignment = center()
    c.border = thin_border()

all_trades = sorted(trades, key=lambda x: x["date"])
for i, t in enumerate(all_trades):
    r = i + 3
    action = t["action"]
    pnl = t.get("pnl", 0) or 0
    cost = t.get("avg_cost", 0) or 0
    ret_pct = pnl / (cost * t.get("shares",1)) if cost and t.get("shares") else 0

    if action == "BUY":
        bg = "DEEAF1"
    elif action == "DIV":
        bg = "E2EFDA"
    else:  # SELL
        bg = C_BULL if pnl > 0 else C_BEAR

    vals = [t["date"], action, t["ticker"], t.get("name",""),
            t.get("shares",0), t.get("price",0), t.get("avg_cost",0),
            pnl, ret_pct if action != "BUY" else None, t.get("reason",""),
            t.get("hold_days",0) if action != "BUY" else None]

    for col, val in enumerate(vals, 1):
        c = ws4.cell(r, col, val)
        c.font = Font(size=9, name="Arial")
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = center() if col in (1,2,3,4,10) else right()

    ws4.cell(r, 5).number_format = "#,##0"
    ws4.cell(r, 6).number_format = "#,##0.00"
    ws4.cell(r, 7).number_format = "#,##0.00"
    c8 = ws4.cell(r, 8)
    c8.number_format = "#,##0;(#,##0);-"
    if action == "SELL":
        c8.font = Font(size=9, name="Arial",
                      color=C_GREEN if pnl > 0 else C_RED, bold=True)
    ws4.cell(r, 9).number_format = "0.0%;(0.0%);-"
    ws4.cell(r, 11).number_format = "#,##0"

set_col_widths(ws4, [12, 8, 10, 12, 8, 12, 12, 12, 10, 32, 10])

# ── Enable autofilter ─────────────────────────────────────────────────────────
ws4.auto_filter.ref = f"A2:K{len(all_trades)+2}"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Per-Ticker Analysis
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Ticker Analysis")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:L1")
ws5["A1"] = "Per-Ticker Trade Analysis"
ws5["A1"].font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
ws5["A1"].fill = fill(C_DARK_BLUE)
ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 28

tk_headers = ["Ticker", "Name", "# Trades", "# Wins", "Win Rate %",
              "Total P&L (฿)", "Avg Win (฿)", "Avg Loss (฿)",
              "Best Trade (฿)", "Worst Trade (฿)", "Avg Hold Days", "Verdict"]
for col, h in enumerate(tk_headers, 1):
    c = ws5.cell(2, col, h)
    c.font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
    c.fill = fill(C_GREY_HDR)
    c.alignment = center()
    c.border = thin_border()

for i, row in df_tk.iterrows():
    r = list(df_tk.index).index(i) + 3
    alt = (r - 3) % 2 == 0
    pnl = row["Total PnL"]
    wr  = row["Win Rate"]
    verdict = ("⭐ Star"    if pnl > 5000 and wr >= 50 else
               "✅ Positive" if pnl > 0 else
               "⚠️ Marginal" if pnl > -3000 else
               "❌ Avoid")

    vals = [row["Ticker"], row["Name"], int(row["Trades"]), int(row["Wins"]),
            row["Win Rate"]/100, row["Total PnL"],
            row["Avg Win"], row["Avg Loss"],
            row["Best Trade"], row["Worst Trade"],
            row["Avg Hold Days"], verdict]

    for col, val in enumerate(vals, 1):
        c = ws5.cell(r, col, val)
        c.font = Font(size=9, name="Arial")
        c.fill = fill(C_GREY_ROW if alt else C_WHITE)
        c.border = thin_border()
        c.alignment = center() if col in (1, 2, 12) else right()

    ws5.cell(r, 5).number_format  = "0.0%"
    ws5.cell(r, 6).number_format  = "#,##0;(#,##0);-"
    ws5.cell(r, 7).number_format  = "#,##0;(#,##0);-"
    ws5.cell(r, 8).number_format  = "#,##0;(#,##0);-"
    ws5.cell(r, 9).number_format  = "#,##0;(#,##0);-"
    ws5.cell(r, 10).number_format = "#,##0;(#,##0);-"
    ws5.cell(r, 11).number_format = "0.0"

    pnl_cell = ws5.cell(r, 6)
    pnl_cell.font = Font(size=9, name="Arial", bold=True,
                        color=C_GREEN if pnl >= 0 else C_RED)

set_col_widths(ws5, [10, 14, 9, 9, 10, 14, 14, 14, 14, 14, 12, 12])
ws5.auto_filter.ref = f"A2:L{len(df_tk)+2}"

# ── Bar chart — P&L by ticker ─────────────────────────────────────────────────
bc = BarChart()
bc.type  = "col"
bc.title = "Total P&L by Ticker"
bc.style = 10
bc.width = 28
bc.height = 14
bc.y_axis.title = "P&L (฿)"
bc.y_axis.numFmt = "#,##0"
bc.x_axis.title = "Ticker"

n_tickers = len(df_tk)
pnl_ref = Reference(ws5, min_col=6, min_row=2, max_row=n_tickers+2)
cats_ref = Reference(ws5, min_col=1, min_row=3, max_row=n_tickers+2)
bc.add_data(pnl_ref, titles_from_data=True)
bc.set_categories(cats_ref)
bc.series[0].graphicalProperties.solidFill = C_MID_BLUE
ws5.add_chart(bc, "N2")

# ── Freeze panes ──────────────────────────────────────────────────────────────
for ws in [ws2, ws4, ws5]:
    ws.freeze_panes = ws.cell(3, 1)
ws3.freeze_panes = ws3.cell(3, 2)

# ── Tab colours ───────────────────────────────────────────────────────────────
wb["Summary"].sheet_properties.tabColor       = C_DARK_BLUE
wb["Equity Curve"].sheet_properties.tabColor  = C_MID_BLUE
wb["Monthly Returns"].sheet_properties.tabColor = C_GREEN
wb["Trade Log"].sheet_properties.tabColor     = C_GREY_HDR
wb["Ticker Analysis"].sheet_properties.tabColor = C_AMBER

wb.save(OUT)
print(f"Saved: {OUT}")
